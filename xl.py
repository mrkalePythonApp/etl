# -*- coding: utf-8 -*-
"""Script for migrating agendas from MS Excel to Family Chronicle."""
__version__ = '0.5.0'
__status__ = 'Beta'
__author__ = 'Libor Gabaj'
__copyright__ = 'Copyright 2019-2020, ' + __author__
__credits__ = [__author__]
__license__ = 'MIT'
__maintainer__ = __author__
__email__ = 'libor.gabaj@gmail.com'

# Standard library modules
import os
import argparse
import logging
from typing import Optional, NoReturn
from importlib import util as imp
import mysql.connector as mysql
import openpyxl

# Custom library modules
import dbconfig as db
import sql
from xl_agenda import Params, Source

# Enumeration and parameter classes
class Script:
    """Script parameters."""
    (fullname, basename, name) = (None, None, None)

class Actuator:
    """Objects of respective processors."""
    (cmdline, logger) = (None, None)


class Target:
    """Parameters of the data target."""
    (
        host, conn, query, cursor, table, database, root, register,
    ) = (None, None, None, None, None, None, None, None,)


# MS Excel actions
def source_open() -> bool:
    """Open a source MS Excel spreadsheet file and return success flag."""
    try:
        Source.wbook = openpyxl.load_workbook(Source.agenda.excel,
                                              data_only=True)
    except Exception as errmsg:
        log = \
            f'Cannot open MS Excel workbook "{Source.agenda.excel}"' \
            f': {errmsg}'
        Actuator.logger.exception(log)
        return False
    return True


###############################################################################
# Migration
###############################################################################
def migrate_sheet() -> bool:
    """Migrate workbook to the target agenda and return success flag."""
    agn = Source.agenda
    agn.reset()
    # Header row - First one with non-empty first column
    for row in Source.wsheet.iter_rows(min_row=1):
        cell = row[0]
        if cell.value:
            agn.header_row = cell.row
            for cellnum, cell in enumerate(row):
                agn.set_column_index(cell.value, cellnum)
            if not agn.check_agenda():
                log = 'Uknown agenda structure'
                Actuator.logger.error(log)
                return False
            break
    if not agn.header_row:
        log = 'No header row detected.'
        Actuator.logger.error(log)
        return False
    # Data rows
    rows = 0
    for row in Source.wsheet.iter_rows(min_row=agn.header_row + 1):
        agn.reset(values_only=True)
        # Process columns of a row
        for cellnum, cell in enumerate(row):
            agn.store_cell(cell, cellnum)
        # Ignore row with empty date column
        if not agn.check_row():
            continue
        # Insert row to target table
        Target.query = sql.compose_insert(
            table=Target.table,
            fields=agn.ins_fields,
            values=agn.ins_values,
            )
        Target.cursor = Target.conn.cursor()
        try:
            Target.cursor.execute(Target.query, agn.dbfields)
            rows += 1
        except mysql.Error as err:
            Actuator.logger.error(err)
    Source.rows += rows
    log = f'Migrated {rows} rows from sheet "{Source.wsheet.title}"'
    Actuator.logger.info(log)
    return True


# Database actions
def connect_db(config: dict) -> Optional[object]:
    """Connect to a database and return connection."""
    try:
        conn = mysql.connect(**config)
        dbname = config['database']
        log = f'Database "{dbname}" connected'
        Actuator.logger.info(log)
        return conn
    except mysql.Error as err:
        log = str(err)
        if err.errno == mysql.errorcode.ER_ACCESS_DENIED_ERROR:
            log = f'Bad database "{dbname}" credentials'
        elif err.errno == mysql.errorcode.ER_BAD_DB_ERROR:
            log = f'Database "{dbname}" does not exist'
        Actuator.logger.error(log)


def target_open() -> bool:
    """Connect to a target database.

    Returns
    -------
    boolean
        Flag about successful processing.

    """
    # Connect to database
    Target.host = db.target_config['host']
    Target.database = db.target_config['database']
    if Target.conn is None:
        Target.conn = connect_db(db.target_config)
        if  Target.conn is None:
            return False
    # Truncate target table
    Target.query = sql.compose_truncate(Target.table)
    Target.cursor = Target.conn.cursor()
    try:
        Target.cursor.execute(Target.query)
        log = f'Table "{Target.database}.{Target.table}" truncated'
        Actuator.logger.info(log)
    except mysql.Error as err:
        log = str(err)
        Actuator.logger.error(log)
        return False
    return True


def target_close():
    """Close cursor and connection to a target database."""
    # Close cursor
    if Target.cursor is not None:
        Target.cursor.close()
    # Close connection to database
    if Target.conn is not None:
        Target.conn.close()
    Target.conn = None
    Target.query = None
    Target.cursor = None
    Target.table = None


###############################################################################
# Setup functions
###############################################################################
def setup_params():
    """Determine script operational parameters."""
    Script.fullname = os.path.splitext(os.path.abspath(__file__))[0]
    Script.basename = os.path.basename(__file__)
    Script.name = os.path.splitext(Script.basename)[0]


def setup_cmdline():
    """Define command line arguments."""
    parser = argparse.ArgumentParser(
        description='Migration of an agenda from provided MS Excel, version '
        + __version__
    )
    # Position arguments
    parser.add_argument(
        'workbook',
        help='Module name without prefix "xl_" and extension' \
            ' for corresponding MS Excel workbook, e.g., "chalupa_events".'
    )
    # Options
    parser.add_argument(
        '-V', '--version',
        action='version',
        version=__version__,
        help='Current version of the script.'
    )
    parser.add_argument(
        '-v', '--verbose',
        choices=['debug', 'info', 'warning', 'error', 'critical'],
        default='debug',
        help='Level of logging to the console.'
    )
    parser.add_argument(
        '-u', '--user',
        type=int,
        default=Params.juser,
        help='Joomla! user id for migration, default: ' + str(Params.juser)
    )
    # Process command line arguments
    Actuator.cmdline = parser.parse_args()


def setup_logger():
    """Configure logging facility."""
    logging.basicConfig(
        level=getattr(logging, Actuator.cmdline.verbose.upper()),
        format='%(levelname)s:%(name)s: %(message)s',
    )
    Actuator.logger = logging.getLogger(Script.name)


def main():
    """Fundamental control function."""
    setup_params()
    setup_cmdline()
    setup_logger()
    # Import plugin module
    module_name = Actuator.cmdline.workbook
    file_path = f'{Script.name}_{module_name}.py'
    try:
        spec = imp.spec_from_file_location(module_name, file_path)
        plugin = imp.module_from_spec(spec)
        spec.loader.exec_module(plugin)
    except Exception as errmsg:
        log = f'Cannot load module "{file_path}": {errmsg}'
        Actuator.logger.exception(log)
        return
    # Connect to MS Excel workbook
    Source.agenda = plugin.workbook()
    Source.agenda.logger = Actuator.logger
    if not source_open():
        return
    # Connect to target database
    Target.table = sql.compose_table(
        sql.target_table_prefix_agenda,
        Source.agenda.agenda)
    if target_open():
        log = \
            f'START -- Migration to database table' \
            f' "{Target.host}//{Target.database}.{Target.table}"'
        Actuator.logger.info(log)
        for Source.wsheet in list(Source.wbook):
            migrate_sheet()
        log = f'STOP -- Migrated {Source.rows} rows in total'
        Actuator.logger.info(log)
    # Close databases
    target_close()


if __name__ == '__main__':
    main()
